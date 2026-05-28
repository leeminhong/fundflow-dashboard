import json
from pathlib import Path

import openpyxl


SOURCE = Path("/Users/minhonglee/Desktop/0528_증감현황.xlsx")
OUTPUT = Path("outputs/fundflow_mvp/source_data.json")


def cell_to_json(value):
    if hasattr(value, "isoformat"):
        return {"type": "date", "value": value.date().isoformat()}
    return value


def main():
    wb_formula = openpyxl.load_workbook(SOURCE, data_only=False)
    wb_values = openpyxl.load_workbook(SOURCE, data_only=True)
    ws_formula = wb_formula["Sheet1"]
    ws_values = wb_values["Sheet1"]

    max_col = 15
    max_row = ws_formula.max_row

    rows = []
    rows_values = []
    for r in range(1, max_row + 1):
        formula_row = []
        value_row = []
        for c in range(1, max_col + 1):
            formula_row.append(cell_to_json(ws_formula.cell(r, c).value))
            value_row.append(cell_to_json(ws_values.cell(r, c).value))
        rows.append(formula_row)
        rows_values.append(value_row)

    headers = [ws_formula.cell(3, c).value for c in range(1, max_col + 1)]
    data_rows = []
    for r in range(4, max_row + 1):
        date_value = ws_values.cell(r, 1).value
        if not date_value:
            continue
        row = {"날짜": date_value.date().isoformat()}
        for idx, header in enumerate(headers[1:], start=2):
            if header:
                row[header] = ws_values.cell(r, idx).value
        data_rows.append(row)

    item_master = [
        {
            "항목코드": "BOND_PUBLIC",
            "섹터": "채권",
            "항목명": "채권 공모",
            "원천잔액컬럼명": "채권 공모",
            "원천증감컬럼명": "채권공모증감",
            "링크": "",
            "표시순서": 10,
            "사용여부": True,
            "단위": "조원",
        },
        {
            "항목코드": "BOND_PRIVATE",
            "섹터": "채권",
            "항목명": "채권 사모",
            "원천잔액컬럼명": "채권 사모",
            "원천증감컬럼명": "채권사모증감",
            "링크": "",
            "표시순서": 20,
            "사용여부": True,
            "단위": "조원",
        },
        {
            "항목코드": "BOND_DISCRETIONARY",
            "섹터": "채권",
            "항목명": "채권투자일임",
            "원천잔액컬럼명": "채권투자일임",
            "원천증감컬럼명": "채권일임증감",
            "링크": "",
            "표시순서": 30,
            "사용여부": True,
            "단위": "조원",
        },
        {
            "항목코드": "EQUITY_PUBLIC",
            "섹터": "주식",
            "항목명": "주식 공모",
            "원천잔액컬럼명": "주식 공모",
            "원천증감컬럼명": "주식공모증감",
            "링크": "",
            "표시순서": 40,
            "사용여부": True,
            "단위": "조원",
        },
        {
            "항목코드": "EQUITY_PRIVATE",
            "섹터": "주식",
            "항목명": "주식 사모",
            "원천잔액컬럼명": "주식 사모",
            "원천증감컬럼명": "주식사모증감",
            "링크": "",
            "표시순서": 50,
            "사용여부": True,
            "단위": "조원",
        },
        {
            "항목코드": "EQUITY_DISCRETIONARY",
            "섹터": "주식",
            "항목명": "주식투자일임",
            "원천잔액컬럼명": "주식투자일임",
            "원천증감컬럼명": "주식일임증감",
            "링크": "",
            "표시순서": 60,
            "사용여부": True,
            "단위": "조원",
        },
        {
            "항목코드": "RP_BALANCE",
            "섹터": "RP",
            "항목명": "RP잔고",
            "원천잔액컬럼명": "RP잔고",
            "원천증감컬럼명": "RP잔고증감",
            "링크": "",
            "표시순서": 70,
            "사용여부": True,
            "단위": "조원",
        },
    ]

    normalized = []
    for raw in data_rows:
        for item in item_master:
            normalized.append(
                {
                    "날짜": raw["날짜"],
                    "섹터": item["섹터"],
                    "항목코드": item["항목코드"],
                    "항목명": item["항목명"],
                    "증감값": raw.get(item["원천증감컬럼명"]),
                    "잔액값": raw.get(item["원천잔액컬럼명"]),
                    "링크": item["링크"],
                    "표시순서": item["표시순서"],
                    "사용여부": item["사용여부"],
                }
            )

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(
        json.dumps(
            {
                "source": str(SOURCE),
                "rawRows": rows,
                "rawValueRows": rows_values,
                "headers": headers,
                "dataRows": data_rows,
                "itemMaster": item_master,
                "normalized": normalized,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
