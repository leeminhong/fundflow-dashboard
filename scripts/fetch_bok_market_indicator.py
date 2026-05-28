#!/usr/bin/env python3
"""Fetch and parse a Bank of Korea daily market-indicator attachment.

The script downloads the balance spreadsheet from a BOK detail page, extracts
the dashboard target items, and can write the web-ready fundflow JSON.
"""

from __future__ import annotations

import argparse
import json
import re
import ssl
import subprocess
from collections import deque
from dataclasses import dataclass
from datetime import date, datetime
from html.parser import HTMLParser
from pathlib import Path
from typing import Iterable
from urllib.parse import parse_qs, unquote, urljoin, urlparse
from urllib.request import Request, urlopen

import openpyxl


USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125 Safari/537.36"
)

BOK_MARKET_LIST_URL = "https://www.bok.or.kr/portal/bbs/P0002018/list.do?menuNo=200366"
BOK_MARKET_RSS_URL = "https://www.bok.or.kr/portal/bbs/P0002018/news.rss?menuNo=200366"
REPO_LINK_URL = (
    "https://seibro.or.kr/websquare/control.jsp?"
    "w2xPath=/IPORTAL/user/repo/BIP_CNTS09001V.xml&menuNo=233"
)
SEIBRO_REPO_FETCH_SCRIPT = Path(__file__).with_name("fetch_seibro_repo.js")
FREESIS_LINK_URL = "https://freesis.kofia.or.kr"
DETAIL_URL_PATTERN = re.compile(
    r"/portal/bbs/P0002018/view\.do\?[^\"'<> ]*nttId=\d+[^\"'<> ]*",
    re.I,
)

ITEM_LINK_OVERRIDES = {
    "FUND_BOND": FREESIS_LINK_URL,
    "FUND_MMF": FREESIS_LINK_URL,
    "FUND_EQUITY": FREESIS_LINK_URL,
    "SEC_CUSTOMER_DEPOSIT": FREESIS_LINK_URL,
    "SEC_CMA": FREESIS_LINK_URL,
    "SEC_CUSTOMER_RP": REPO_LINK_URL,
    "REPO_INTERBANK": REPO_LINK_URL,
}

LEGACY_FUND_ITEM_CODES = {"FUND_BOND", "FUND_MMF", "FUND_EQUITY"}

FREESIS_SUMMARY_COLUMNS = {
    "FUND_BOND_PUBLIC_DOMESTIC": "공모_국내_채권",
    "FUND_BOND_PRIVATE_DOMESTIC": "사모_국내_채권",
    "FUND_BOND_DISCRETIONARY_DOMESTIC": "일임_국내_채권",
    "FUND_MMF_PUBLIC_DOMESTIC": "공모_국내_MMF",
    "FUND_MMF_PRIVATE_DOMESTIC": "사모_국내_MMF",
    "FUND_EQUITY_PUBLIC_DOMESTIC": "공모_국내_주식",
    "FUND_EQUITY_PUBLIC_OVERSEAS": "공모_해외_주식",
    "FUND_EQUITY_PRIVATE_DOMESTIC": "사모_국내_주식",
    "FUND_EQUITY_PRIVATE_OVERSEAS": "사모_해외_주식",
    "FUND_EQUITY_DISCRETIONARY_DOMESTIC": "일임_국내_주식",
    "FUND_EQUITY_DISCRETIONARY_OVERSEAS": "일임_해외_주식",
}

FREESIS_CALCULATED_PARENTS = [
    {
        "itemCode": "FUND_MMF_TOTAL", "itemName": "MMF", "displayOrder": 40,
        "children": ["FUND_MMF_PUBLIC_DOMESTIC", "FUND_MMF_PRIVATE_DOMESTIC"],
    },
    {
        "itemCode": "FUND_BOND_TOTAL", "itemName": "채권", "displayOrder": 45,
        "children": ["FUND_BOND_PUBLIC_DOMESTIC", "FUND_BOND_PRIVATE_DOMESTIC", "FUND_BOND_DISCRETIONARY_DOMESTIC"],
    },
    {
        "itemCode": "FUND_EQUITY_TOTAL", "itemName": "주식", "displayOrder": 55,
        "children": [
            "FUND_EQUITY_PUBLIC_DOMESTIC", "FUND_EQUITY_PUBLIC_OVERSEAS",
            "FUND_EQUITY_PRIVATE_DOMESTIC", "FUND_EQUITY_PRIVATE_OVERSEAS",
            "FUND_EQUITY_DISCRETIONARY_DOMESTIC", "FUND_EQUITY_DISCRETIONARY_OVERSEAS",
        ],
    },
]

FREESIS_FUND_ITEMS = [
    {"itemCode": "FUND_MMF_PUBLIC_DOMESTIC", "itemName": "공모 국내 MMF", "displayOrder": 41, "parentCode": "FUND_MMF_TOTAL"},
    {"itemCode": "FUND_MMF_PRIVATE_DOMESTIC", "itemName": "사모 국내 MMF", "displayOrder": 42, "parentCode": "FUND_MMF_TOTAL"},
    {"itemCode": "FUND_BOND_PUBLIC_DOMESTIC", "itemName": "공모 국내 채권", "displayOrder": 46, "parentCode": "FUND_BOND_TOTAL"},
    {"itemCode": "FUND_BOND_PRIVATE_DOMESTIC", "itemName": "사모 국내 채권", "displayOrder": 47, "parentCode": "FUND_BOND_TOTAL"},
    {"itemCode": "FUND_BOND_DISCRETIONARY_DOMESTIC", "itemName": "일임 국내 채권", "displayOrder": 48, "parentCode": "FUND_BOND_TOTAL"},
    {"itemCode": "FUND_EQUITY_PUBLIC_DOMESTIC", "itemName": "공모 국내 주식", "displayOrder": 56, "parentCode": "FUND_EQUITY_TOTAL"},
    {"itemCode": "FUND_EQUITY_PUBLIC_OVERSEAS", "itemName": "공모 해외 주식", "displayOrder": 57, "parentCode": "FUND_EQUITY_TOTAL"},
    {"itemCode": "FUND_EQUITY_PRIVATE_DOMESTIC", "itemName": "사모 국내 주식", "displayOrder": 58, "parentCode": "FUND_EQUITY_TOTAL"},
    {"itemCode": "FUND_EQUITY_PRIVATE_OVERSEAS", "itemName": "사모 해외 주식", "displayOrder": 59, "parentCode": "FUND_EQUITY_TOTAL"},
    {"itemCode": "FUND_EQUITY_DISCRETIONARY_DOMESTIC", "itemName": "일임 국내 주식", "displayOrder": 60, "parentCode": "FUND_EQUITY_TOTAL"},
    {"itemCode": "FUND_EQUITY_DISCRETIONARY_OVERSEAS", "itemName": "일임 해외 주식", "displayOrder": 61, "parentCode": "FUND_EQUITY_TOTAL"},
]


@dataclass(frozen=True)
class Attachment:
    label: str
    url: str


@dataclass(frozen=True)
class PageResult:
    page_url: str
    page_title: str
    attachment: Attachment
    parsed: dict


TARGET_ITEMS = {
    "실세총예금": {
        "sector": "은행",
        "itemCode": "BANK_TOTAL_DEPOSIT",
        "itemName": "실세총예금",
        "displayOrder": 10,
    },
    "실세요구불": {
        "sector": "은행",
        "itemCode": "BANK_DEMAND_DEPOSIT",
        "itemName": "실세요구불",
        "parentCode": "BANK_TOTAL_DEPOSIT",
        "level": 2,
        "includeInTotal": False,
        "displayOrder": 11,
    },
    "저축성": {
        "sector": "은행",
        "itemCode": "BANK_SAVINGS_DEPOSIT",
        "itemName": "저축성",
        "parentCode": "BANK_TOTAL_DEPOSIT",
        "level": 2,
        "includeInTotal": False,
        "displayOrder": 12,
    },
    "금전신탁": {"sector": "은행", "itemCode": "BANK_MONEY_TRUST", "itemName": "금전신탁", "displayOrder": 20},
    "고객예탁금": {"sector": "증권", "itemCode": "SEC_CUSTOMER_DEPOSIT", "itemName": "고객예탁금", "displayOrder": 70},
    "대고객RP매도": {"sector": "증권", "itemCode": "SEC_CUSTOMER_RP", "itemName": "고객RP", "displayOrder": 80},
    "CMA": {"sector": "증권", "itemCode": "SEC_CMA", "itemName": "CMA", "displayOrder": 90},
    "채권형": {"sector": "투신", "itemCode": "FUND_BOND", "itemName": "채권형", "displayOrder": 40},
    "MMF": {"sector": "투신", "itemCode": "FUND_MMF", "itemName": "MMF", "displayOrder": 50},
    "주식형": {"sector": "투신", "itemCode": "FUND_EQUITY", "itemName": "주식형", "displayOrder": 60},
}

SECTOR_ORDER = {"은행": 1, "REPO": 2, "투신": 3, "증권": 4}

SEIBRO_REPO_ITEM = {
    "itemCode": "REPO_INTERBANK",
    "sector": "REPO",
    "groupName": "기관RP",
    "itemName": "기관RP",
    "parentCode": None,
    "level": 1,
    "itemType": "raw",
    "includeInTotal": True,
    "requiredForComplete": True,
    "showInHeatmap": True,
    "rawBalanceColumn": "잔고금액",
    "rawChangeColumn": "잔고금액_증감",
    "link": REPO_LINK_URL,
    "displayOrder": 65,
    "isActive": True,
    "unit": "조원",
    "source": "SEIBro 일별거래현황",
}


class LinkParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.links: list[tuple[str, str]] = []
        self._current_href: str | None = None
        self._text_parts: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() != "a":
            return
        attrs_dict = dict(attrs)
        href = attrs_dict.get("href")
        if href:
            self._current_href = href
            self._text_parts = []

    def handle_data(self, data: str) -> None:
        if self._current_href:
            self._text_parts.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() == "a" and self._current_href:
            label = " ".join("".join(self._text_parts).split())
            self.links.append((self._current_href, label))
            self._current_href = None
            self._text_parts = []


def fetch_bytes(url: str) -> bytes:
    context = ssl._create_unverified_context()
    req = Request(url, headers={"User-Agent": USER_AGENT})
    with urlopen(req, timeout=30, context=context) as response:
        return response.read()


def fetch_text(url: str) -> str:
    return fetch_bytes(url).decode("utf-8", "replace")


def extract_title(html: str) -> str:
    meta_match = re.search(
        r'<meta\s+property=["\']title["\']\s+content=["\']([^"\']+)["\']',
        html,
        re.I,
    )
    if meta_match:
        return meta_match.group(1).strip()

    match = re.search(r"<h[12][^>]*>\s*(.*?)\s*</h[12]>", html, re.I | re.S)
    if match:
        return re.sub(r"\s+", " ", re.sub(r"<[^>]+>", "", match.group(1))).strip()
    match = re.search(r"<title>\s*(.*?)\s*</title>", html, re.I | re.S)
    if match:
        return re.sub(r"\s+", " ", match.group(1)).strip()
    return "(title not found)"


def extract_attachments(page_url: str, html: str) -> list[Attachment]:
    parser = LinkParser()
    parser.feed(html)

    by_url: dict[str, Attachment] = {}
    for href, label in parser.links:
        text = unquote(" ".join([href, label]))
        if not re.search(r"\.(xlsx|xls)\b", text, re.I):
            continue
        file_label = label or Path(unquote(href).split("?")[0]).name
        absolute_url = urljoin(page_url, href)
        current = by_url.get(absolute_url)
        if current is None or re.search(r"\.(xlsx|xls)\b", file_label, re.I):
            by_url[absolute_url] = Attachment(label=file_label, url=absolute_url)

    return list(by_url.values())


def choose_balance_attachment(attachments: Iterable[Attachment]) -> Attachment:
    candidates = list(attachments)
    if not candidates:
        raise RuntimeError("No Excel attachments found on the page.")

    balance = [item for item in candidates if "잔액" in unquote(item.label)]
    if balance:
        return balance[0]
    return candidates[0]


def extract_ntt_id(url: str) -> int | None:
    parsed = urlparse(url)
    raw = parse_qs(parsed.query).get("nttId", [None])[0]
    if raw is None:
        return None
    try:
        return int(raw)
    except ValueError:
        return None


def canonicalize_detail_url(url: str) -> str | None:
    ntt_id = extract_ntt_id(url)
    if ntt_id is None:
        return None
    return f"https://www.bok.or.kr/portal/bbs/P0002018/view.do?nttId={ntt_id}&menuNo=200366"


def extract_related_page_urls(page_url: str, html: str) -> list[str]:
    related = []
    seen = set()
    for match in DETAIL_URL_PATTERN.findall(html):
        absolute = urljoin(page_url, match)
        canonical = canonicalize_detail_url(absolute)
        if canonical and canonical not in seen:
            seen.add(canonical)
            related.append(canonical)
    return related


def discover_recent_page_urls(seed_url: str, count: int, max_pages: int = 24) -> tuple[list[str], dict[str, str]]:
    seed = canonicalize_detail_url(seed_url)
    if seed is None:
        raise RuntimeError("Seed URL does not contain nttId.")

    queue = deque([seed])
    html_cache: dict[str, str] = {}

    while queue and len(html_cache) < max_pages:
        current_url = queue.popleft()
        if current_url in html_cache:
            continue
        try:
            html = fetch_text(current_url)
        except Exception as exc:  # pragma: no cover - network boundary
            print(f"warning: failed to fetch {current_url}: {exc}")
            continue

        html_cache[current_url] = html
        for related in extract_related_page_urls(current_url, html):
            if related not in html_cache and related not in queue:
                queue.append(related)

        if len(html_cache) >= max(count + 3, count * 2):
            break

    sorted_urls = sorted(
        html_cache.keys(),
        key=lambda url: extract_ntt_id(url) or 0,
        reverse=True,
    )
    return sorted_urls[:count], html_cache


def extract_rss_page_urls(limit: int = 40) -> list[str]:
    rss = fetch_text(BOK_MARKET_RSS_URL)
    pattern = re.compile(
        r"<link><!\[CDATA\[(https://www\.bok\.or\.kr/portal/bbs/P0002018/view\.do\?[^]]+)\]\]></link>",
        re.I,
    )
    urls = []
    seen = set()
    for raw in pattern.findall(rss):
        canonical = canonicalize_detail_url(raw)
        if canonical and canonical not in seen:
            seen.add(canonical)
            urls.append(canonical)
        if len(urls) >= limit:
            break
    return urls


def safe_filename(name: str) -> str:
    cleaned = unquote(name).strip().replace("/", "_")
    return cleaned or "bok_market_indicator.xlsx"


def normalize_label(value: object) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"\d+\)", "", text)
    text = re.sub(r"[\s()]+", "", text)
    return text.replace("*", "")


def excel_date(value: object) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return None


def to_number(value: object) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if value is None:
        return None
    text = str(value).strip().replace(",", "")
    if text in {"", "-", ".."}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_ymd(value: object) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y%m%d"):
        try:
            return datetime.strptime(text[:10], fmt).date().isoformat()
        except ValueError:
            continue
    return None


def fetch_seibro_repo_rows(limit: int) -> list[dict]:
    if not SEIBRO_REPO_FETCH_SCRIPT.exists():
        raise FileNotFoundError(f"SEIBro fetch script not found: {SEIBRO_REPO_FETCH_SCRIPT}")

    cmd = ["node", str(SEIBRO_REPO_FETCH_SCRIPT), "--limit", str(limit)]
    proc = subprocess.run(cmd, check=True, capture_output=True, text=True)
    raw = proc.stdout.strip()
    if not raw:
        return []
    rows = json.loads(raw)
    parsed = []
    for row in rows:
        date_iso = parse_ymd(row.get("date"))
        balance_billion = to_number(row.get("balanceAmountBillion"))
        trade_billion = to_number(row.get("tradeAmountBillion"))
        if not date_iso or balance_billion is None:
            continue
        parsed.append(
            {
                "date": date_iso,
                "tradeAmountBillion": trade_billion,
                "balanceAmountBillion": balance_billion,
            }
        )
    parsed.sort(key=lambda row: row["date"])
    return parsed


def apply_seibro_repo(data: dict, rows: list[dict]) -> None:
    if not rows:
        return

    item = next((row for row in data["items"] if row["itemCode"] == SEIBRO_REPO_ITEM["itemCode"]), None)
    if item is None:
        item = dict(SEIBRO_REPO_ITEM)
        data["items"].append(item)

    data["records"] = [record for record in data["records"] if record["itemCode"] != item["itemCode"]]

    prev_balance = None
    for row in rows:
        balance = round(row["balanceAmountBillion"] / 100, 4)
        change = 0.0 if prev_balance is None else round(balance - prev_balance, 4)
        prev_balance = balance
        data["records"].append(
            {
                "date": row["date"],
                "sector": item["sector"],
                "groupName": item["itemName"],
                "itemCode": item["itemCode"],
                "itemName": item["itemName"],
                "parentCode": item.get("parentCode"),
                "level": item.get("level", 1),
                "itemType": item.get("itemType", "raw"),
                "includeInTotal": item.get("includeInTotal", True),
                "requiredForComplete": item.get("requiredForComplete", True),
                "showInHeatmap": item.get("showInHeatmap", True),
                "changeValue": change,
                "balanceValue": balance,
                "link": REPO_LINK_URL,
                "displayOrder": item["displayOrder"],
                "isActive": item.get("isActive", True),
                "hasSourceMapping": True,
                "source": "SEIBro 일별거래현황",
                "sourceLabel": "잔고금액",
                "sourceUnit": "십억원",
                "changeDate": row["date"],
                "balanceDate": row["date"],
                "sourcePageUrl": REPO_LINK_URL,
                "sourcePageTitle": "SEIBro Repo 시장현황",
                "sourceAttachment": "",
                "sourceAttachmentUrl": "",
            }
        )

    data["meta"]["seibroRepoRows"] = len(rows)
    data["meta"]["seibroRepoLatestDate"] = rows[-1]["date"]


def parse_market_workbook(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_name = "일일동향" if "일일동향" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    section_row = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "2. 금융권별 여수신 동향":
                section_row = cell.row
                break
        if section_row:
            break
    if section_row is None:
        raise RuntimeError("Could not find '2. 금융권별 여수신 동향' section.")

    header_row = section_row + 3
    latest_change_col = None
    latest_change_date = None
    balance_col = None
    balance_date = None

    for col in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col).value
        parsed_date = excel_date(value)
        header_label = normalize_label(ws.cell(header_row - 1, col).value)
        if parsed_date and header_label == "잔액":
            balance_col = col
            balance_date = parsed_date
            break

    if balance_col is not None:
        for col in range(1, balance_col):
            parsed_date = excel_date(ws.cell(header_row, col).value)
            if parsed_date:
                latest_change_col = col
                latest_change_date = parsed_date

    if latest_change_col is None or balance_col is None:
        raise RuntimeError("Could not locate latest change and balance date columns.")

    found: dict[str, dict] = {}
    for row_idx in range(section_row + 1, ws.max_row + 1):
        raw_label = ws.cell(row_idx, 3).value
        normalized = normalize_label(raw_label)
        if normalized not in TARGET_ITEMS:
            continue

        meta = TARGET_ITEMS[normalized]
        change_okrw = to_number(ws.cell(row_idx, latest_change_col).value)
        balance_okrw = to_number(ws.cell(row_idx, balance_col).value)
        found[normalized] = {
            **meta,
            "sourceLabel": str(raw_label).strip(),
            "date": balance_date,
            "changeDate": latest_change_date,
            "balanceDate": balance_date,
            "changeValueOkrw": change_okrw,
            "balanceValueOkrw": balance_okrw,
            "changeValueTrillionKrw": None if change_okrw is None else change_okrw / 10000,
            "balanceValueTrillionKrw": None if balance_okrw is None else balance_okrw / 10000,
            "unit": "조원",
            "sourceUnit": "억원",
        }

    missing = [
        TARGET_ITEMS[key]["itemName"]
        for key in TARGET_ITEMS
        if key not in found
    ]
    records = [found[key] for key in TARGET_ITEMS if key in found]
    return {
        "sourceFile": str(path),
        "sheetName": sheet_name,
        "latestChangeDate": latest_change_date,
        "balanceDate": balance_date,
        "records": records,
        "missingItems": missing,
    }


def recompute_status_summary(data: dict) -> None:
    items = sorted(data["items"], key=lambda item: item["displayOrder"])
    active_items = [item for item in items if item.get("isActive", True)]
    active_codes = [item["itemCode"] for item in active_items]
    data["items"] = items
    data["records"] = sorted(
        data["records"],
        key=lambda record: (record["date"], record["displayOrder"], record["itemCode"]),
    )
    data["dates"] = sorted({record["date"] for record in data["records"]})
    data["sectors"] = sorted({item["sector"] for item in active_items}, key=lambda sector: SECTOR_ORDER.get(sector, 99))

    date_status = []
    complete_dates = []
    records_by_date: dict[str, dict[str, dict]] = {}
    for record in data["records"]:
        if not record.get("isActive", True):
            continue
        records_by_date.setdefault(record["date"], {})[record["itemCode"]] = record

    for date_value in data["dates"]:
        per_code = records_by_date.get(date_value, {})
        missing = []
        for item in active_items:
            if item.get("itemType") == "calculated":
                continue
            row = per_code.get(item["itemCode"])
            if row is None or row.get("changeValue") is None or row.get("balanceValue") is None:
                missing.append(item["itemName"])
        is_complete = len(missing) == 0
        if is_complete:
            complete_dates.append(date_value)
        date_status.append(
            {
                "date": date_value,
                "isComplete": is_complete,
                "missingItems": missing,
                "pendingItems": missing,
                "filledItemCount": len(active_codes) - len(missing),
                "totalItemCount": len(active_codes),
            }
        )

    data["dateStatus"] = date_status
    latest_date = data["dates"][-1] if data["dates"] else None
    default_date = complete_dates[-1] if complete_dates else latest_date

    default_records = [
        record
        for record in data["records"]
        if record["date"] == default_date and record.get("includeInTotal", True) and record.get("changeValue") is not None
    ]
    sector_summary = []
    for sector in data["sectors"]:
        sector_records = [record for record in default_records if record["sector"] == sector]
        sector_summary.append(
            {
                "sector": sector,
                "latestChange": sum(record["changeValue"] or 0 for record in sector_records),
                "latestBalance": sum(record["balanceValue"] or 0 for record in sector_records),
                "itemCount": len(sector_records),
            }
        )

    data["summary"] = {
        "latestDate": latest_date,
        "defaultDate": default_date,
        "totalLatestChange": sum(record["changeValue"] or 0 for record in default_records),
        "sectorSummary": sector_summary,
    }
    data["meta"]["latestDate"] = latest_date
    data["meta"]["defaultDate"] = default_date
    data["meta"]["historyWindow"] = len(data["dates"])


def apply_freesis_stock_deposit(data: dict, wb, freesis_summary_path: Path) -> None:
    if "투자자예탁금" not in wb.sheetnames:
        return

    ws = wb["투자자예탁금"]
    rows = list(ws.values)
    if len(rows) < 2:
        return

    headers = [str(v).strip() if v is not None else "" for v in rows[0]]
    header_idx = {name: idx for idx, name in enumerate(headers)}
    deposit_col = next((h for h in headers if "투자자예탁금" in h), None)
    if not deposit_col:
        return

    data["items"] = [item for item in data["items"] if item["itemCode"] != "SEC_CUSTOMER_DEPOSIT"]
    data["records"] = [record for record in data["records"] if record["itemCode"] != "SEC_CUSTOMER_DEPOSIT"]

    data["items"].append(
        {
            "itemCode": "SEC_CUSTOMER_DEPOSIT",
            "sector": "증권",
            "groupName": "고객예탁금",
            "itemName": "고객예탁금",
            "parentCode": None,
            "level": 1,
            "itemType": "raw",
            "includeInTotal": True,
            "requiredForComplete": True,
            "showInHeatmap": True,
            "rawBalanceColumn": deposit_col,
            "rawChangeColumn": f"{deposit_col}_증감",
            "link": FREESIS_LINK_URL,
            "displayOrder": 70,
            "isActive": True,
            "unit": "조원",
            "source": "FREESIS 증시자금추이",
        }
    )

    date_rows = []
    for raw in rows[1:]:
        date_iso = parse_ymd(raw[header_idx["기준일자"]])
        if not date_iso:
            continue
        date_rows.append((date_iso, raw))
    date_rows.sort(key=lambda item: item[0])

    prev_balance = None
    for date_iso, raw in date_rows:
        raw_value = raw[header_idx[deposit_col]]
        balance_million = to_number(raw_value)
        balance = None if balance_million is None else balance_million / 1000000
        change = None
        if balance is not None and prev_balance is not None:
            change = balance - prev_balance
        elif balance is not None:
            change = 0.0
        if balance is not None:
            prev_balance = balance

        data["records"].append(
            {
                "date": date_iso,
                "sector": "증권",
                "groupName": "고객예탁금",
                "itemCode": "SEC_CUSTOMER_DEPOSIT",
                "itemName": "고객예탁금",
                "parentCode": None,
                "level": 1,
                "itemType": "raw",
                "includeInTotal": True,
                "requiredForComplete": True,
                "showInHeatmap": True,
                "changeValue": change,
                "balanceValue": balance,
                "link": FREESIS_LINK_URL,
                "displayOrder": 70,
                "isActive": True,
                "hasSourceMapping": True,
                "source": "FREESIS 증시자금추이",
                "sourceLabel": deposit_col,
                "sourceUnit": "백만원",
                "changeDate": date_iso,
                "balanceDate": date_iso,
                "sourcePageUrl": FREESIS_LINK_URL,
                "sourcePageTitle": "FREESIS 증시자금추이",
                "sourceAttachment": freesis_summary_path.name,
                "sourceAttachmentUrl": "",
            }
        )


def apply_freesis_summary(data: dict, freesis_summary_path: Path) -> None:
    wb = openpyxl.load_workbook(freesis_summary_path, data_only=True)

    summary_sheet = None
    for name in ("펀드일임_요약", "요약"):
        if name in wb.sheetnames:
            summary_sheet = name
            break
    if summary_sheet is None:
        raise RuntimeError(f"'펀드일임_요약' or '요약' sheet not found in {freesis_summary_path}")
    ws = wb[summary_sheet]
    rows = list(ws.values)
    if not rows:
        raise RuntimeError(f"'{summary_sheet}' sheet is empty in {freesis_summary_path}")

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    header_idx = {name: idx for idx, name in enumerate(headers)}
    available_fund_items = [
        meta for meta in FREESIS_FUND_ITEMS
        if FREESIS_SUMMARY_COLUMNS[meta["itemCode"]] in header_idx
    ]
    if "기준일자" not in header_idx:
        raise RuntimeError(f"'기준일자' column not found in {summary_sheet} sheet")

    # Remove legacy fund items/records from BOK aggregate before injecting FREESIS detail rows.
    data["items"] = [item for item in data["items"] if item["itemCode"] not in LEGACY_FUND_ITEM_CODES]
    data["records"] = [record for record in data["records"] if record["itemCode"] not in LEGACY_FUND_ITEM_CODES]

    # Add calculated parent items (MMF, 채권, 주식)
    for parent in FREESIS_CALCULATED_PARENTS:
        available_children = [meta["itemCode"] for meta in available_fund_items]
        if any(child in available_children for child in parent["children"]):
            data["items"].append(
                {
                    "itemCode": parent["itemCode"],
                    "sector": "투신",
                    "groupName": parent["itemName"],
                    "itemName": parent["itemName"],
                    "parentCode": None,
                    "level": 1,
                    "itemType": "calculated",
                    "includeInTotal": True,
                    "requiredForComplete": False,
                    "showInHeatmap": True,
                    "rawBalanceColumn": None,
                    "rawChangeColumn": None,
                    "link": FREESIS_LINK_URL,
                    "displayOrder": parent["displayOrder"],
                    "isActive": True,
                    "unit": "조원",
                    "source": "FREESIS 유형별 설정/일임",
                }
            )

    # Add child items
    for meta in available_fund_items:
        data["items"].append(
            {
                "itemCode": meta["itemCode"],
                "sector": "투신",
                "groupName": meta.get("parentCode", "투신"),
                "itemName": meta["itemName"],
                "parentCode": meta.get("parentCode"),
                "level": 2 if meta.get("parentCode") else 1,
                "itemType": "raw",
                "includeInTotal": False,
                "requiredForComplete": True,
                "showInHeatmap": True,
                "rawBalanceColumn": FREESIS_SUMMARY_COLUMNS[meta["itemCode"]],
                "rawChangeColumn": f"{FREESIS_SUMMARY_COLUMNS[meta['itemCode']]}_증감",
                "link": FREESIS_LINK_URL,
                "displayOrder": meta["displayOrder"],
                "isActive": True,
                "unit": "조원",
                "source": "FREESIS 유형별 설정/일임",
            }
        )

    # Date-keyed rows to calculate change values from consecutive available observations.
    date_rows = []
    for raw in rows[1:]:
        date_iso = parse_ymd(raw[header_idx["기준일자"]])
        if not date_iso:
            continue
        date_rows.append((date_iso, raw))
    date_rows.sort(key=lambda item: item[0])

    prev_balance: dict[str, float] = {}
    for date_iso, raw in date_rows:
        for meta in available_fund_items:
            item_code = meta["itemCode"]
            col_name = FREESIS_SUMMARY_COLUMNS[item_code]
            raw_value = raw[header_idx[col_name]]
            balance_okrw = to_number(raw_value)
            balance = None if balance_okrw is None else balance_okrw / 10000
            change = None
            if balance is not None and item_code in prev_balance:
                change = balance - prev_balance[item_code]
            elif balance is not None:
                change = 0.0
                prev_balance[item_code] = balance

            data["records"].append(
                {
                    "date": date_iso,
                    "sector": "투신",
                    "groupName": meta.get("parentCode", "투신"),
                    "itemCode": item_code,
                    "itemName": meta["itemName"],
                    "parentCode": meta.get("parentCode"),
                    "level": 2 if meta.get("parentCode") else 1,
                    "itemType": "raw",
                    "includeInTotal": False,
                    "requiredForComplete": True,
                    "showInHeatmap": True,
                    "changeValue": change,
                    "balanceValue": balance,
                    "link": FREESIS_LINK_URL,
                    "displayOrder": meta["displayOrder"],
                    "isActive": True,
                    "hasSourceMapping": True,
                    "source": "FREESIS 유형별 설정/일임",
                    "sourceLabel": col_name,
                    "sourceUnit": "억원",
                    "changeDate": date_iso,
                    "balanceDate": date_iso,
                    "sourcePageUrl": FREESIS_LINK_URL,
                    "sourcePageTitle": "FREESIS 유형별 기간설정",
                    "sourceAttachment": freesis_summary_path.name,
                    "sourceAttachmentUrl": "",
                }
            )

    # Compute calculated parent records (MMF, 채권, 주식) by summing children per date.
    all_dates = sorted({r["date"] for r in data["records"] if r["sector"] == "투신"})
    records_by_date_code: dict[tuple[str, str], dict] = {}
    for record in data["records"]:
        records_by_date_code[(record["date"], record["itemCode"])] = record

    for parent in FREESIS_CALCULATED_PARENTS:
        for date_iso in all_dates:
            child_records = [
                records_by_date_code[(date_iso, child_code)]
                for child_code in parent["children"]
                if (date_iso, child_code) in records_by_date_code
            ]
            if not child_records:
                continue
            change_values = [r["changeValue"] for r in child_records if r["changeValue"] is not None]
            balance_values = [r["balanceValue"] for r in child_records if r["balanceValue"] is not None]
            data["records"].append(
                {
                    "date": date_iso,
                    "sector": "투신",
                    "groupName": parent["itemName"],
                    "itemCode": parent["itemCode"],
                    "itemName": parent["itemName"],
                    "parentCode": None,
                    "level": 1,
                    "itemType": "calculated",
                    "includeInTotal": True,
                    "requiredForComplete": False,
                    "showInHeatmap": True,
                    "changeValue": sum(change_values) if change_values else None,
                    "balanceValue": sum(balance_values) if balance_values else None,
                    "link": FREESIS_LINK_URL,
                    "displayOrder": parent["displayOrder"],
                    "isActive": True,
                    "hasSourceMapping": False,
                    "source": "FREESIS 유형별 설정/일임 (합산)",
                }
            )

    apply_freesis_stock_deposit(data, wb, freesis_summary_path)

    data["meta"]["freesisSummaryFile"] = str(freesis_summary_path)
    recompute_status_summary(data)


def item_link(item_code: str) -> str:
    return ITEM_LINK_OVERRIDES.get(item_code, BOK_MARKET_LIST_URL)


def parse_page_result(page_url: str, html: str, output_dir: Path) -> PageResult:
    title = extract_title(html)
    attachments = extract_attachments(page_url, html)
    selected = choose_balance_attachment(attachments)

    output_path = output_dir / safe_filename(selected.label)
    output_path.write_bytes(fetch_bytes(selected.url))
    parsed = parse_market_workbook(output_path)
    parsed_path = output_path.with_suffix(".json")
    parsed_path.write_text(json.dumps(parsed, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"page_title: {title}")
    print(f"page_url: {page_url}")
    print(f"attachments_found: {len(attachments)}")
    print(f"selected_attachment: {selected.label}")
    print(f"downloaded: {output_path}")
    print(f"bytes: {output_path.stat().st_size}")
    print(f"parsed_json: {parsed_path}")
    print(f"balance_date: {parsed['balanceDate']}")
    if parsed["missingItems"]:
        print("missing_items: " + ", ".join(parsed["missingItems"]))

    return PageResult(
        page_url=page_url,
        page_title=title,
        attachment=selected,
        parsed=parsed,
    )


def build_web_data(page_results: list[PageResult]) -> dict:
    if not page_results:
        raise RuntimeError("No page results to build web data.")

    latest_by_item: dict[str, dict] = {}
    records_by_key: dict[tuple[str, str], dict] = {}
    records = []
    item_map: dict[str, dict] = {}

    for source_key, source_meta in TARGET_ITEMS.items():
        item = {
            "itemCode": source_meta["itemCode"],
            "sector": source_meta["sector"],
            "groupName": source_meta["itemName"],
            "itemName": source_meta["itemName"],
            "parentCode": source_meta.get("parentCode"),
            "level": source_meta.get("level", 1),
            "itemType": "raw",
            "includeInTotal": source_meta.get("includeInTotal", True),
            "requiredForComplete": True,
            "showInHeatmap": True,
            "rawBalanceColumn": source_key,
            "rawChangeColumn": source_key,
            "link": item_link(source_meta["itemCode"]),
            "displayOrder": source_meta["displayOrder"],
            "isActive": True,
            "unit": "조원",
            "source": "한국은행 일일 금융시장 주요지표",
        }
        item_map[item["itemCode"]] = item

    for page_result in sorted(
        page_results,
        key=lambda row: extract_ntt_id(row.page_url) or 0,
        reverse=True,
    ):
        for record in page_result.parsed["records"]:
            item = item_map[record["itemCode"]]
            key = (record["balanceDate"], record["itemCode"])
            if key in records_by_key:
                continue

            payload = {
                "date": record["balanceDate"],
                "sector": record["sector"],
                "groupName": record["itemName"],
                "itemCode": record["itemCode"],
                "itemName": record["itemName"],
                "parentCode": item["parentCode"],
                "level": item["level"],
                "itemType": "raw",
                "includeInTotal": item["includeInTotal"],
                "requiredForComplete": True,
                "showInHeatmap": True,
                "changeValue": record["changeValueTrillionKrw"],
                "balanceValue": record["balanceValueTrillionKrw"],
                "link": item["link"],
                "displayOrder": item["displayOrder"],
                "isActive": True,
                "hasSourceMapping": True,
                "source": "한국은행 일일 금융시장 주요지표",
                "sourceLabel": record["sourceLabel"],
                "sourceUnit": record["sourceUnit"],
                "changeDate": record["changeDate"],
                "balanceDate": record["balanceDate"],
                "sourcePageUrl": page_result.page_url,
                "sourcePageTitle": page_result.page_title,
                "sourceAttachment": page_result.attachment.label,
                "sourceAttachmentUrl": page_result.attachment.url,
            }
            records_by_key[key] = payload

            previous = latest_by_item.get(record["itemCode"])
            if previous is None or (previous["date"] < payload["date"]):
                latest_by_item[record["itemCode"]] = payload

    items = sorted(item_map.values(), key=lambda item: item["displayOrder"])
    records = sorted(records_by_key.values(), key=lambda record: (record["date"], record["displayOrder"]))
    sectors = sorted({item["sector"] for item in items}, key=lambda sector: SECTOR_ORDER.get(sector, 99))
    dates = sorted({record["date"] for record in records})
    if not dates:
        raise RuntimeError("No valid dates were parsed from market indicator files.")

    date_status = []
    default_date = dates[-1]
    complete_dates = []
    expected_codes = [item["itemCode"] for item in items if item["isActive"]]

    for date_value in dates:
        date_records = [record for record in records if record["date"] == date_value and record["isActive"]]
        by_code = {record["itemCode"]: record for record in date_records}
        missing_items = []
        for item in items:
            target = by_code.get(item["itemCode"])
            if target is None:
                missing_items.append(item["itemName"])
                continue
            if target["changeValue"] is None or target["balanceValue"] is None:
                missing_items.append(item["itemName"])

        is_complete = len(missing_items) == 0
        if is_complete:
            complete_dates.append(date_value)

        date_status.append(
            {
                "date": date_value,
                "isComplete": is_complete,
                "missingItems": missing_items,
                "pendingItems": missing_items,
                "filledItemCount": len(expected_codes) - len(missing_items),
                "totalItemCount": len(expected_codes),
            }
        )

    if complete_dates:
        default_date = complete_dates[-1]

    default_records = [
        record
        for record in records
        if record["date"] == default_date and record["changeValue"] is not None
        and record["includeInTotal"]
    ]
    sector_summary = []
    for sector in sectors:
        sector_records = [record for record in default_records if record["sector"] == sector]
        sector_summary.append(
            {
                "sector": sector,
                "latestChange": sum(record["changeValue"] or 0 for record in sector_records),
                "latestBalance": sum(record["balanceValue"] or 0 for record in sector_records),
                "itemCount": len(sector_records),
            }
        )

    return {
        "meta": {
            "title": "Daily Fundflow Dashboard",
            "sourceFile": ", ".join(sorted({row.parsed["sourceFile"] for row in page_results})),
            "sourcePageTitle": page_results[0].page_title,
            "sourcePageUrl": page_results[0].page_url,
            "sourceAttachment": page_results[0].attachment.label,
            "sourceAttachmentUrl": page_results[0].attachment.url,
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "latestDate": dates[-1],
            "defaultDate": default_date,
            "unit": "조원",
            "version": 3,
            "historyWindow": len(dates),
        },
        "items": items,
        "dates": dates,
        "sectors": sectors,
        "dateStatus": date_status,
        "summary": {
            "latestDate": dates[-1],
            "defaultDate": default_date,
            "totalLatestChange": sum(record["changeValue"] or 0 for record in default_records),
            "sectorSummary": sector_summary,
        },
        "records": records,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "url",
        help="BOK detail page URL, e.g. https://www.bok.or.kr/portal/bbs/P0002018/view.do?nttId=...&menuNo=200366",
    )
    parser.add_argument(
        "--output-dir",
        default="outputs/bok_market_indicator",
        help="Directory to save the downloaded attachment.",
    )
    parser.add_argument(
        "--write-web-data",
        default=None,
        help="Optional path to write dashboard data JSON, e.g. data/fundflow.json.",
    )
    parser.add_argument(
        "--days",
        type=int,
        default=1,
        help="How many recent market-indicator pages to aggregate (default: 1).",
    )
    parser.add_argument(
        "--freesis-summary-xlsx",
        default=None,
        help="Optional FREESIS summary xlsx path (요약 sheet) to replace fund items with 11 detailed rows.",
    )
    parser.add_argument(
        "--skip-seibro-repo",
        action="store_true",
        help="Skip SEIBro Repo 잔고금액 merge for SEC_CUSTOMER_RP.",
    )
    parser.add_argument(
        "--seibro-repo-limit",
        type=int,
        default=7,
        help="How many recent SEIBro daily rows to merge (default: 7).",
    )
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    page_urls, html_cache = discover_recent_page_urls(args.url, args.days)
    rss_urls = extract_rss_page_urls(limit=max(args.days * 6, 20))
    candidate_urls = sorted(
        set(page_urls + rss_urls),
        key=lambda url: extract_ntt_id(url) or 0,
        reverse=True,
    )

    print(f"seed_url: {args.url}")
    print(f"discovered_pages_from_related: {len(page_urls)}")
    print(f"discovered_pages_from_rss: {len(rss_urls)}")
    print(f"candidate_pages: {len(candidate_urls)}")
    for idx, page_url in enumerate(candidate_urls[:20], start=1):
        print(f"[{idx}] {page_url}")

    page_results = []
    seen_balance_dates = set()
    for page_url in candidate_urls:
        if len(page_results) >= args.days:
            break
        html = html_cache.get(page_url) or fetch_text(page_url)
        try:
            page_result = parse_page_result(page_url, html, output_dir)
        except Exception as exc:
            print(f"skip_page: {page_url} ({exc})")
            continue

        balance_date = page_result.parsed["balanceDate"]
        if balance_date in seen_balance_dates:
            print(f"skip_duplicate_date: {page_url} ({balance_date})")
            continue
        seen_balance_dates.add(balance_date)

        page_results.append(page_result)
        print("parsed_records:")
        for record in page_result.parsed["records"]:
            change = record["changeValueTrillionKrw"]
            balance = record["balanceValueTrillionKrw"]
            change_text = "-" if change is None else f"{change:,.4f}"
            balance_text = "-" if balance is None else f"{balance:,.4f}"
            print(
                f"- {record['sector']} / {record['itemName']}: "
                f"증감 {change_text}조원, 잔액 {balance_text}조원"
            )
    print(f"selected_pages: {len(page_results)}")
    if len(page_results) < args.days:
        print(f"warning: requested {args.days} pages but only {len(page_results)} valid pages were found.")

    if args.write_web_data:
        web_data = build_web_data(page_results)
        if args.freesis_summary_xlsx:
            freesis_path = Path(args.freesis_summary_xlsx)
            if not freesis_path.exists():
                raise FileNotFoundError(f"FREESIS summary file not found: {freesis_path}")
            apply_freesis_summary(web_data, freesis_path)
            print(f"freesis_summary_applied: {freesis_path}")
        if not args.skip_seibro_repo:
            try:
                seibro_rows = fetch_seibro_repo_rows(args.seibro_repo_limit)
                apply_seibro_repo(web_data, seibro_rows)
                recompute_status_summary(web_data)
                if seibro_rows:
                    print(f"seibro_repo_applied: {len(seibro_rows)} rows (latest {seibro_rows[-1]['date']})")
                else:
                    print("seibro_repo_applied: 0 rows")
            except Exception as exc:
                print(f"warning: seibro_repo_merge_failed: {exc}")
        web_data_path = Path(args.write_web_data)
        web_data_path.parent.mkdir(parents=True, exist_ok=True)
        web_data_path.write_text(json.dumps(web_data, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"web_data: {web_data_path}")


if __name__ == "__main__":
    main()
