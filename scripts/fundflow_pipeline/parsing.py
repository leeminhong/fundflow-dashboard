"""Value/date normalization and BOK workbook parsing."""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path

import openpyxl

from .config import TARGET_ITEMS

def normalize_label(value: object) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"\d+\)", "", text)
    text = re.sub(r"[\s()]+", "", text)
    return text.replace("*", "")


def excel_date(value: object) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return None


def to_number(value: object) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if value is None:
        return None
    text = str(value).strip().replace(",", "")
    if text in {"", "-", ".."}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_ymd(value: object) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y%m%d"):
        try:
            return datetime.strptime(text[:10], fmt).date().isoformat()
        except ValueError:
            continue
    return None


# 월말 "잔액" 백필 게시물은 날짜별로 시트가 나뉜다(예: '5.13','5.12',...,'4.30').
# 일반 일일 리포트는 '일일동향' 단일 시트.
DATE_SHEET_RE = re.compile(r"^\s*\d{1,2}\.\d{1,2}\s*$")


def _parse_market_sheet(ws) -> dict | None:
    """Parse a single market-indicator worksheet.

    Returns a dict with records/balanceDate, or None when the sheet does not
    contain a parseable '2. 금융권별 여수신 동향' section (e.g. cover/summary
    sheets in a multi-sheet backfill workbook).
    """
    section_row = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "2. 금융권별 여수신 동향":
                section_row = cell.row
                break
        if section_row:
            break
    if section_row is None:
        return None

    header_row = section_row + 3
    latest_change_col = None
    latest_change_date = None
    balance_col = None
    balance_date = None

    for col in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col).value
        parsed_date = excel_date(value)
        header_label = normalize_label(ws.cell(header_row - 1, col).value)
        if parsed_date and header_label == "잔액":
            balance_col = col
            balance_date = parsed_date
            break

    if balance_col is not None:
        for col in range(1, balance_col):
            parsed_date = excel_date(ws.cell(header_row, col).value)
            if parsed_date:
                latest_change_col = col
                latest_change_date = parsed_date

    if latest_change_col is None or balance_col is None:
        return None

    found: dict[str, dict] = {}
    for row_idx in range(section_row + 1, ws.max_row + 1):
        raw_label = ws.cell(row_idx, 3).value
        normalized = normalize_label(raw_label)
        if normalized not in TARGET_ITEMS:
            continue

        meta = TARGET_ITEMS[normalized]
        change_okrw = to_number(ws.cell(row_idx, latest_change_col).value)
        balance_okrw = to_number(ws.cell(row_idx, balance_col).value)
        found[normalized] = {
            **meta,
            "sourceLabel": str(raw_label).strip(),
            "date": balance_date,
            "changeDate": latest_change_date,
            "balanceDate": balance_date,
            "changeValueOkrw": change_okrw,
            "balanceValueOkrw": balance_okrw,
            "changeValueTrillionKrw": None if change_okrw is None else change_okrw / 10000,
            "balanceValueTrillionKrw": None if balance_okrw is None else balance_okrw / 10000,
            "unit": "조원",
            "sourceUnit": "억원",
        }

    missing = [
        TARGET_ITEMS[key]["itemName"]
        for key in TARGET_ITEMS
        if key not in found
    ]
    records = [found[key] for key in TARGET_ITEMS if key in found]
    return {
        "sheetName": ws.title,
        "latestChangeDate": latest_change_date,
        "balanceDate": balance_date,
        "records": records,
        "missingItems": missing,
    }


def parse_market_workbook(path: Path) -> dict:
    """Parse a BOK market-indicator workbook.

    Daily reports have a single '일일동향' sheet (one balance date). Month-end
    "잔액" backfill posts instead carry one date-named sheet per missing date
    (e.g. '5.13' … '4.30'); every such sheet is parsed and its records are
    merged so the previously-missing dates get filled. Each record carries its
    own balanceDate/changeDate, so downstream merge keys them per date.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    if "일일동향" in wb.sheetnames:
        target_sheets = ["일일동향"]
    else:
        # Backfill workbook: parse every date-named sheet (fall back to the
        # first sheet if none match the date pattern).
        target_sheets = [name for name in wb.sheetnames if DATE_SHEET_RE.match(name)]
        if not target_sheets:
            target_sheets = [wb.sheetnames[0]]

    parsed_sheets = []
    for name in target_sheets:
        result = _parse_market_sheet(wb[name])
        if result and result["balanceDate"]:
            parsed_sheets.append(result)

    if not parsed_sheets:
        raise RuntimeError(
            "Could not parse any '2. 금융권별 여수신 동향' section "
            f"(sheets tried: {target_sheets})."
        )

    all_records = []
    for sheet in parsed_sheets:
        all_records.extend(sheet["records"])

    # Top-level date fields describe the newest sheet (used for logging and the
    # cli's per-page dedup); the full date set lives in balanceDates.
    latest = max(parsed_sheets, key=lambda sheet: sheet["balanceDate"])
    return {
        "sourceFile": str(path),
        "sheetName": ", ".join(sheet["sheetName"] for sheet in parsed_sheets),
        "latestChangeDate": latest["latestChangeDate"],
        "balanceDate": latest["balanceDate"],
        "balanceDates": sorted({sheet["balanceDate"] for sheet in parsed_sheets}),
        "records": all_records,
        "missingItems": latest["missingItems"],
    }


