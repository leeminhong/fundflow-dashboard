"""FREESIS deposit/fund data injection from the cumulative DB."""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl

from .config import (
    FREESIS_CALCULATED_PARENTS,
    FREESIS_FUND_ITEMS,
    FREESIS_LINK_URL,
    FREESIS_SUMMARY_COLUMNS,
    LEGACY_FUND_ITEM_CODES,
    REPO_LINK_URL,
    SEIBRO_REPO_ITEM,
)
from .parsing import parse_ymd, to_number
from .webdata import recompute_status_summary

def apply_freesis_stock_deposit(data: dict, wb, freesis_summary_path: Path) -> None:
    if "투자자예탁금" not in wb.sheetnames:
        return

    ws = wb["투자자예탁금"]
    rows = list(ws.values)
    if len(rows) < 2:
        return

    headers = [str(v).strip() if v is not None else "" for v in rows[0]]
    header_idx = {name: idx for idx, name in enumerate(headers)}
    deposit_col = next((h for h in headers if "투자자예탁금" in h), None)
    if not deposit_col:
        return

    data["items"] = [item for item in data["items"] if item["itemCode"] != "SEC_CUSTOMER_DEPOSIT"]
    data["records"] = [record for record in data["records"] if record["itemCode"] != "SEC_CUSTOMER_DEPOSIT"]

    data["items"].append(
        {
            "itemCode": "SEC_CUSTOMER_DEPOSIT",
            "sector": "증권",
            "groupName": "고객예탁금",
            "itemName": "고객예탁금",
            "parentCode": None,
            "level": 1,
            "itemType": "raw",
            "includeInTotal": True,
            "requiredForComplete": True,
            "showInHeatmap": True,
            "rawBalanceColumn": deposit_col,
            "rawChangeColumn": f"{deposit_col}_증감",
            "link": FREESIS_LINK_URL,
            "displayOrder": 70,
            "isActive": True,
            "unit": "조원",
            "source": "FREESIS 증시자금추이",
        }
    )

    date_rows = []
    for raw in rows[1:]:
        date_iso = parse_ymd(raw[header_idx["기준일자"]])
        if not date_iso:
            continue
        date_rows.append((date_iso, raw))
    date_rows.sort(key=lambda item: item[0])

    prev_balance = None
    for date_iso, raw in date_rows:
        raw_value = raw[header_idx[deposit_col]]
        balance_million = to_number(raw_value)
        balance = None if balance_million is None else balance_million / 1000000
        change = None
        if balance is not None and prev_balance is not None:
            change = balance - prev_balance
        elif balance is not None:
            change = 0.0
        if balance is not None:
            prev_balance = balance

        data["records"].append(
            {
                "date": date_iso,
                "sector": "증권",
                "groupName": "고객예탁금",
                "itemCode": "SEC_CUSTOMER_DEPOSIT",
                "itemName": "고객예탁금",
                "parentCode": None,
                "level": 1,
                "itemType": "raw",
                "includeInTotal": True,
                "requiredForComplete": True,
                "showInHeatmap": True,
                "changeValue": change,
                "balanceValue": balance,
                "link": FREESIS_LINK_URL,
                "displayOrder": 70,
                "isActive": True,
                "hasSourceMapping": True,
                "source": "FREESIS 증시자금추이",
                "sourceLabel": deposit_col,
                "sourceUnit": "백만원",
                "changeDate": date_iso,
                "balanceDate": date_iso,
                "sourcePageUrl": FREESIS_LINK_URL,
                "sourcePageTitle": "FREESIS 증시자금추이",
                "sourceAttachment": freesis_summary_path.name,
                "sourceAttachmentUrl": "",
            }
        )


def apply_freesis_summary(data: dict, freesis_summary_path: Path) -> None:
    wb = openpyxl.load_workbook(freesis_summary_path, data_only=True)

    summary_sheet = None
    for name in ("펀드일임_요약", "요약"):
        if name in wb.sheetnames:
            summary_sheet = name
            break
    if summary_sheet is None:
        raise RuntimeError(f"'펀드일임_요약' or '요약' sheet not found in {freesis_summary_path}")
    ws = wb[summary_sheet]
    rows = list(ws.values)
    if not rows:
        raise RuntimeError(f"'{summary_sheet}' sheet is empty in {freesis_summary_path}")

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    header_idx = {name: idx for idx, name in enumerate(headers)}
    available_fund_items = [
        meta for meta in FREESIS_FUND_ITEMS
        if FREESIS_SUMMARY_COLUMNS[meta["itemCode"]] in header_idx
    ]
    if "기준일자" not in header_idx:
        raise RuntimeError(f"'기준일자' column not found in {summary_sheet} sheet")

    # Remove legacy fund items/records from BOK aggregate before injecting FREESIS detail rows.
    data["items"] = [item for item in data["items"] if item["itemCode"] not in LEGACY_FUND_ITEM_CODES]
    data["records"] = [record for record in data["records"] if record["itemCode"] not in LEGACY_FUND_ITEM_CODES]

    # Add calculated parent items (MMF, 채권, 주식)
    for parent in FREESIS_CALCULATED_PARENTS:
        available_children = [meta["itemCode"] for meta in available_fund_items]
        if any(child in available_children for child in parent["children"]):
            data["items"].append(
                {
                    "itemCode": parent["itemCode"],
                    "sector": "투신",
                    "groupName": parent["itemName"],
                    "itemName": parent["itemName"],
                    "parentCode": None,
                    "level": 1,
                    "itemType": "calculated",
                    "includeInTotal": True,
                    "requiredForComplete": False,
                    "showInHeatmap": True,
                    "rawBalanceColumn": None,
                    "rawChangeColumn": None,
                    "link": FREESIS_LINK_URL,
                    "displayOrder": parent["displayOrder"],
                    "isActive": True,
                    "unit": "조원",
                    "source": "FREESIS 유형별 설정/일임",
                }
            )

    # Add child items
    for meta in available_fund_items:
        data["items"].append(
            {
                "itemCode": meta["itemCode"],
                "sector": "투신",
                "groupName": meta.get("parentCode", "투신"),
                "itemName": meta["itemName"],
                "parentCode": meta.get("parentCode"),
                "level": 2 if meta.get("parentCode") else 1,
                "itemType": "raw",
                "includeInTotal": False,
                "requiredForComplete": True,
                "showInHeatmap": True,
                "rawBalanceColumn": FREESIS_SUMMARY_COLUMNS[meta["itemCode"]],
                "rawChangeColumn": f"{FREESIS_SUMMARY_COLUMNS[meta['itemCode']]}_증감",
                "link": FREESIS_LINK_URL,
                "displayOrder": meta["displayOrder"],
                "isActive": True,
                "unit": "조원",
                "source": "FREESIS 유형별 설정/일임",
            }
        )

    # Date-keyed rows to calculate change values from consecutive available observations.
    date_rows = []
    for raw in rows[1:]:
        date_iso = parse_ymd(raw[header_idx["기준일자"]])
        if not date_iso:
            continue
        date_rows.append((date_iso, raw))
    date_rows.sort(key=lambda item: item[0])

    prev_balance: dict[str, float] = {}
    for date_iso, raw in date_rows:
        for meta in available_fund_items:
            item_code = meta["itemCode"]
            col_name = FREESIS_SUMMARY_COLUMNS[item_code]
            raw_value = raw[header_idx[col_name]]
            balance_okrw = to_number(raw_value)
            balance = None if balance_okrw is None else balance_okrw / 10000
            change = None
            if balance is not None and item_code in prev_balance:
                change = balance - prev_balance[item_code]
            elif balance is not None:
                change = 0.0
                prev_balance[item_code] = balance

            data["records"].append(
                {
                    "date": date_iso,
                    "sector": "투신",
                    "groupName": meta.get("parentCode", "투신"),
                    "itemCode": item_code,
                    "itemName": meta["itemName"],
                    "parentCode": meta.get("parentCode"),
                    "level": 2 if meta.get("parentCode") else 1,
                    "itemType": "raw",
                    "includeInTotal": False,
                    "requiredForComplete": True,
                    "showInHeatmap": True,
                    "changeValue": change,
                    "balanceValue": balance,
                    "link": FREESIS_LINK_URL,
                    "displayOrder": meta["displayOrder"],
                    "isActive": True,
                    "hasSourceMapping": True,
                    "source": "FREESIS 유형별 설정/일임",
                    "sourceLabel": col_name,
                    "sourceUnit": "억원",
                    "changeDate": date_iso,
                    "balanceDate": date_iso,
                    "sourcePageUrl": FREESIS_LINK_URL,
                    "sourcePageTitle": "FREESIS 유형별 기간설정",
                    "sourceAttachment": freesis_summary_path.name,
                    "sourceAttachmentUrl": "",
                }
            )

    # Compute calculated parent records (MMF, 채권, 주식) by summing children per date.
    all_dates = sorted({r["date"] for r in data["records"] if r["sector"] == "투신"})
    records_by_date_code: dict[tuple[str, str], dict] = {}
    for record in data["records"]:
        records_by_date_code[(record["date"], record["itemCode"])] = record

    for parent in FREESIS_CALCULATED_PARENTS:
        for date_iso in all_dates:
            child_records = [
                records_by_date_code[(date_iso, child_code)]
                for child_code in parent["children"]
                if (date_iso, child_code) in records_by_date_code
            ]
            if not child_records:
                continue
            change_values = [r["changeValue"] for r in child_records if r["changeValue"] is not None]
            balance_values = [r["balanceValue"] for r in child_records if r["balanceValue"] is not None]
            data["records"].append(
                {
                    "date": date_iso,
                    "sector": "투신",
                    "groupName": parent["itemName"],
                    "itemCode": parent["itemCode"],
                    "itemName": parent["itemName"],
                    "parentCode": None,
                    "level": 1,
                    "itemType": "calculated",
                    "includeInTotal": True,
                    "requiredForComplete": False,
                    "showInHeatmap": True,
                    "changeValue": sum(change_values) if change_values else None,
                    "balanceValue": sum(balance_values) if balance_values else None,
                    "link": FREESIS_LINK_URL,
                    "displayOrder": parent["displayOrder"],
                    "isActive": True,
                    "hasSourceMapping": False,
                    "source": "FREESIS 유형별 설정/일임 (합산)",
                }
            )

    apply_freesis_stock_deposit(data, wb, freesis_summary_path)

    data["meta"]["freesisSummaryFile"] = str(freesis_summary_path)
    recompute_status_summary(data)


def apply_freesis_db(data: dict, db_path: Path) -> None:
    """Read freesis_db.json and inject fund + stock-deposit records."""
    db = json.loads(db_path.read_text(encoding="utf-8"))

    fund_summary = db.get("fundSummary", {})
    stock_deposit = db.get("stockDeposit", {})
    if not fund_summary and not stock_deposit:
        print(f"freesis_db: no data in {db_path}")
        return

    # --- Fund items ---
    # Remove legacy BOK aggregate fund items
    data["items"] = [i for i in data["items"] if i["itemCode"] not in LEGACY_FUND_ITEM_CODES]
    data["records"] = [r for r in data["records"] if r["itemCode"] not in LEGACY_FUND_ITEM_CODES]

    # Determine which child items are available from DB
    sample_cols = next(iter(fund_summary.values()), {})
    available_fund_items = [
        meta for meta in FREESIS_FUND_ITEMS
        if FREESIS_SUMMARY_COLUMNS[meta["itemCode"]] in sample_cols
    ]

    # Add calculated parent items
    for parent in FREESIS_CALCULATED_PARENTS:
        available_codes = {m["itemCode"] for m in available_fund_items}
        if any(c in available_codes for c in parent["children"]):
            data["items"].append({
                "itemCode": parent["itemCode"], "sector": "투신",
                "groupName": parent["itemName"], "itemName": parent["itemName"],
                "parentCode": None, "level": 1, "itemType": "calculated",
                "includeInTotal": True, "requiredForComplete": False,
                "showInHeatmap": True, "rawBalanceColumn": None,
                "rawChangeColumn": None, "link": FREESIS_LINK_URL,
                "displayOrder": parent["displayOrder"], "isActive": True,
                "unit": "조원", "source": "FREESIS 유형별 설정/일임",
            })

    # Add child items
    for meta in available_fund_items:
        data["items"].append({
            "itemCode": meta["itemCode"], "sector": "투신",
            "groupName": meta.get("parentCode", "투신"),
            "itemName": meta["itemName"],
            "parentCode": meta.get("parentCode"),
            "level": 2 if meta.get("parentCode") else 1,
            "itemType": "raw", "includeInTotal": False,
            "requiredForComplete": True, "showInHeatmap": True,
            "rawBalanceColumn": FREESIS_SUMMARY_COLUMNS[meta["itemCode"]],
            "rawChangeColumn": f"{FREESIS_SUMMARY_COLUMNS[meta['itemCode']]}_증감",
            "link": FREESIS_LINK_URL, "displayOrder": meta["displayOrder"],
            "isActive": True, "unit": "조원",
            "source": "FREESIS 유형별 설정/일임",
        })

    # Add fund summary records
    fund_changes = db.get("fundChanges", {})
    sorted_dates = sorted(fund_summary.keys())
    for date_key in sorted_dates:
        row = fund_summary[date_key]
        date_iso = parse_ymd(date_key)
        if not date_iso:
            continue
        change_row = fund_changes.get(date_key, {})
        for meta in available_fund_items:
            col_name = FREESIS_SUMMARY_COLUMNS[meta["itemCode"]]
            raw_val = to_number(row.get(col_name))
            balance = None if raw_val is None else raw_val / 10000
            raw_change = to_number(change_row.get(col_name))
            change = None if raw_change is None else raw_change / 10000
            data["records"].append({
                "date": date_iso, "sector": "투신",
                "groupName": meta.get("parentCode", "투신"),
                "itemCode": meta["itemCode"], "itemName": meta["itemName"],
                "parentCode": meta.get("parentCode"),
                "level": 2 if meta.get("parentCode") else 1,
                "itemType": "raw", "includeInTotal": False,
                "requiredForComplete": True, "showInHeatmap": True,
                "changeValue": change, "balanceValue": balance,
                "link": FREESIS_LINK_URL, "displayOrder": meta["displayOrder"],
                "isActive": True, "hasSourceMapping": True,
                "source": "FREESIS 유형별 설정/일임",
                "sourceLabel": col_name, "sourceUnit": "억원",
                "changeDate": date_iso, "balanceDate": date_iso,
                "sourcePageUrl": FREESIS_LINK_URL,
                "sourcePageTitle": "FREESIS 유형별 기간설정",
                "sourceAttachment": db_path.name, "sourceAttachmentUrl": "",
            })

    # Compute calculated parent records
    all_dates = sorted({r["date"] for r in data["records"] if r["sector"] == "투신"})
    rec_map: dict[tuple[str, str], dict] = {}
    for r in data["records"]:
        rec_map[(r["date"], r["itemCode"])] = r
    for parent in FREESIS_CALCULATED_PARENTS:
        for date_iso in all_dates:
            children = [rec_map[(date_iso, c)] for c in parent["children"] if (date_iso, c) in rec_map]
            if not children:
                continue
            cv = [r["changeValue"] for r in children if r["changeValue"] is not None]
            bv = [r["balanceValue"] for r in children if r["balanceValue"] is not None]
            data["records"].append({
                "date": date_iso, "sector": "투신",
                "groupName": parent["itemName"],
                "itemCode": parent["itemCode"],
                "itemName": parent["itemName"],
                "parentCode": None, "level": 1,
                "itemType": "calculated", "includeInTotal": True,
                "requiredForComplete": False, "showInHeatmap": True,
                "changeValue": sum(cv) if cv else None,
                "balanceValue": sum(bv) if bv else None,
                "link": FREESIS_LINK_URL,
                "displayOrder": parent["displayOrder"], "isActive": True,
                "hasSourceMapping": False,
                "source": "FREESIS 유형별 설정/일임 (합산)",
            })

    # --- Stock deposit (투자자예탁금) ---
    stock_changes = db.get("stockDepositChanges", {})
    if stock_deposit:
        # Remove existing BOK-sourced deposit
        data["items"] = [i for i in data["items"] if i["itemCode"] != "SEC_CUSTOMER_DEPOSIT"]
        data["records"] = [r for r in data["records"] if r["itemCode"] != "SEC_CUSTOMER_DEPOSIT"]

        data["items"].append({
            "itemCode": "SEC_CUSTOMER_DEPOSIT", "sector": "증권",
            "groupName": "고객예탁금", "itemName": "고객예탁금",
            "parentCode": None, "level": 1, "itemType": "raw",
            "includeInTotal": True, "requiredForComplete": True,
            "showInHeatmap": True, "rawBalanceColumn": "투자자예탁금",
            "rawChangeColumn": "투자자예탁금_증감",
            "link": FREESIS_LINK_URL, "displayOrder": 70, "isActive": True,
            "unit": "조원", "source": "FREESIS 증시자금추이",
        })

        sorted_dep = sorted(stock_deposit.items())
        for date_key, raw_val in sorted_dep:
            date_iso = parse_ymd(date_key)
            if not date_iso:
                continue
            val = to_number(raw_val)
            balance = None if val is None else val / 1000000  # 백만원 → 조원
            raw_change = to_number(stock_changes.get(date_key))
            change = None if raw_change is None else raw_change / 1000000
            data["records"].append({
                "date": date_iso, "sector": "증권",
                "groupName": "고객예탁금",
                "itemCode": "SEC_CUSTOMER_DEPOSIT",
                "itemName": "고객예탁금",
                "parentCode": None, "level": 1, "itemType": "raw",
                "includeInTotal": True, "requiredForComplete": True,
                "showInHeatmap": True,
                "changeValue": change, "balanceValue": balance,
                "link": FREESIS_LINK_URL, "displayOrder": 70,
                "isActive": True, "hasSourceMapping": True,
                "source": "FREESIS 증시자금추이",
                "sourceLabel": "투자자예탁금", "sourceUnit": "백만원",
                "changeDate": date_iso, "balanceDate": date_iso,
                "sourcePageUrl": FREESIS_LINK_URL,
                "sourcePageTitle": "FREESIS 증시자금추이",
                "sourceAttachment": db_path.name, "sourceAttachmentUrl": "",
            })

    # --- Repo balance (기관RP) from DB ---
    repo_balance = db.get("repoBalance", {})
    repo_changes = db.get("repoChanges", {})
    if repo_balance:
        # Remove existing SEIBro-sourced repo records
        data["items"] = [i for i in data["items"] if i["itemCode"] != "REPO_INTERBANK"]
        data["records"] = [r for r in data["records"] if r["itemCode"] != "REPO_INTERBANK"]

        data["items"].append(dict(SEIBRO_REPO_ITEM))

        for date_key in sorted(repo_balance.keys()):
            date_iso = parse_ymd(date_key)
            if not date_iso:
                continue
            balance = to_number(repo_balance[date_key])
            raw_change = to_number(repo_changes.get(date_key))
            change = raw_change  # 이미 조원 단위
            if balance is None:
                continue
            data["records"].append({
                "date": date_iso, "sector": "REPO",
                "groupName": "기관RP", "itemCode": "REPO_INTERBANK",
                "itemName": "기관RP",
                "parentCode": None, "level": 1, "itemType": "raw",
                "includeInTotal": True, "requiredForComplete": True,
                "showInHeatmap": True,
                "changeValue": change, "balanceValue": balance,
                "link": REPO_LINK_URL, "displayOrder": 65,
                "isActive": True, "hasSourceMapping": True,
                "source": "SEIBro Repo 시장현황",
                "sourceLabel": "잔고금액", "sourceUnit": "조원",
                "changeDate": date_iso, "balanceDate": date_iso,
                "sourcePageUrl": REPO_LINK_URL,
                "sourcePageTitle": "SEIBro Repo 시장현황",
                "sourceAttachment": db_path.name, "sourceAttachmentUrl": "",
            })

    data["meta"]["freesisDbFile"] = str(db_path)
    data["meta"]["freesisDbLastUpdated"] = db.get("lastUpdated")
    recompute_status_summary(data)
    print(f"freesis_db_applied: {db_path} "
          f"(fund {len(fund_summary)} days, deposit {len(stock_deposit)} days, "
          f"repo {len(repo_balance)} days)")


